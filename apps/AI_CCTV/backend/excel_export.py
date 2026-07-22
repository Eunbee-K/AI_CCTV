from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from .frames import _img_bytes_downscaled
from .rows import build_results_view
from .state import state


def export_excel(path: str) -> Optional[str]:
    """엑셀 보고서를 path에 저장. 실패 시 에러 메시지를 반환, 성공 시 None."""
    wb = Workbook()
    ws = wb.active
    ws.title = "CCTV 조사 집계표"

    font_bold = Font(name='맑은 고딕', size=12, bold=True)
    font_norm = Font(name='맑은 고딕', size=12)
    align_c = Alignment(horizontal='center', vertical='center', wrap_text=True)

    thick = Side(style='thick')
    thin = Side(style='thin')
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
    border_bottom = Border(bottom=thin)
    gray_fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    ws['B3'] = ' ▣ 공 사 명 :'
    ws['B3'].font = font_bold
    ws['B4'] = ' ▣ 관로번호 :'
    ws['B4'].font = font_bold
    ws.merge_cells('C3:H3')
    ws['C3'] = state.site_name
    ws['C3'].font = font_norm

    pid_val = ""
    if state.video_data_map:
        pid_val = list(state.video_data_map.values())[0].get("pipe_id", "")
    ws.merge_cells('C4:H4')
    ws['C4'] = pid_val
    ws['C4'].font = font_norm

    for c in range(1, 19):
        ws.cell(row=4, column=c).border = border_bottom

    widths = {'A': 2, 'B': 15, 'C': 9, 'D': 14, 'E': 10, 'F': 10, 'G': 21, 'H': 13}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    current_row = 5
    H_PT = 28.8
    IMG_WIDTH = 740

    display = build_results_view()
    for item in display:
        if item["type"] == "separator":
            continue
        if item["type"] == "group":
            for child in item["children"]:
                if child.get("fp"):  # 오탐으로 마킹된 행은 보고서에서 제외
                    continue
                _export_excel_row(ws, child, current_row, H_PT, IMG_WIDTH,
                                   thick, border_all, gray_fill, align_c,
                                   font_bold, font_norm)
                current_row += 16
        else:
            if item.get("fp"):
                continue
            _export_excel_row(ws, item, current_row, H_PT, IMG_WIDTH,
                               thick, border_all, gray_fill, align_c,
                               font_bold, font_norm)
            current_row += 16

    try:
        wb.save(path)
        return None
    except Exception as e:
        return str(e)


def _export_excel_row(ws, row, current_row, H_PT, IMG_WIDTH,
                       thick, border_all, gray_fill, align_c,
                       font_bold, font_norm):
    fname = row["filename"]
    t = row["time_s"]

    direction = row.get("direction", "")
    fp = None
    v_data = state.video_data_map.get(fname)
    if v_data:
        data_row = next((r for r in v_data["rows"] if r["time"] == t), None)
        if data_row:
            # bbox+신뢰도가 그려진 프레임이 있으면 보고서에도 그걸 사용
            fp = data_row.get("frame_annot_path") or data_row.get("frame_path")

    r_start = current_row + 1
    r_end = current_row + 11
    ws.merge_cells(f'B{r_start}:H{r_end}')
    img_cell = ws[f'B{r_start}']

    for r in range(r_start, r_end + 1):
        ws.cell(r, 2).border = Border(left=thick)
        ws.cell(r, 8).border = Border(right=thick)
    for rx in range(r_start, r_end + 1):
        ws.row_dimensions[rx].height = 28.3

    if fp and fp.exists():
        try:
            d, _, _ = _img_bytes_downscaled(fp, 80, fixed_width=IMG_WIDTH)
            img = XLImage(BytesIO(d))
            ws.add_image(img, f'B{r_start}')
        except Exception:
            img_cell.value = "Image Error"

    h_row = current_row + 13
    headers = ["주행방향", "영상위치", "관로번호", "관경", "거리", "결함종류", "비고"]
    for idx, txt in enumerate(headers, 2):
        c = ws.cell(h_row, idx, txt)
        c.font = font_bold
        c.alignment = align_c
        c.fill = gray_fill
        c.border = border_all
    ws.row_dimensions[h_row].height = H_PT

    d_row = current_row + 14
    pid_row = v_data["pipe_id"] if v_data else ""
    dia_row = v_data["dia"] if v_data else ""

    row_vals = [
        direction, row["time_str"], pid_row, dia_row,
        row.get("dist", ""), ", ".join(row.get("defects", [])), row.get("note", ""),
    ]
    for idx, v in enumerate(row_vals, 2):
        c = ws.cell(d_row, idx, v)
        c.font = font_norm
        c.alignment = align_c
        c.border = border_all
    ws.row_dimensions[d_row].height = H_PT
