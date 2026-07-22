"""run별 결과를 CSV 요약표에 누적하고, 메일 발송용 엑셀 파일을 만든다."""
import csv
from pathlib import Path

FIELDNAMES = [
    "sweep_name", "run_name", "status",
    "imgsz", "epochs", "batch", "model",
    "train_images", "val_images", "test_images",
    "best_epoch", "best_precision", "best_recall",
    "best_map50", "best_map50_95",
    "final_map50", "final_map50_95",
    "started_at", "finished_at",
]


def parse_ultralytics_results_csv(results_csv_path) -> dict:
    """ultralytics train()이 남기는 results.csv에서 best/최종 epoch 지표를 뽑는다.

    "best"는 metrics/mAP50-95(B) 기준으로 가장 높은 epoch을 고른다.
    (experiments/*/metadata.yaml에 기존에 수기로 기록해온 방식과 동일한 기준)
    """
    results_csv_path = Path(results_csv_path)
    if not results_csv_path.exists():
        return {}

    with open(results_csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        rows = [{k.strip(): v.strip() for k, v in row.items()} for row in reader]

    if not rows:
        return {}

    def col(row, *candidates):
        for c in candidates:
            if c in row and row[c] not in ("", None):
                return row[c]
        return None

    def as_float(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    map5095_key_candidates = ("metrics/mAP50-95(B)", "metrics/mAP50-95")
    map50_key_candidates = ("metrics/mAP50(B)", "metrics/mAP50")
    precision_candidates = ("metrics/precision(B)", "metrics/precision")
    recall_candidates = ("metrics/recall(B)", "metrics/recall")
    epoch_candidates = ("epoch",)

    scored = [(as_float(col(r, *map5095_key_candidates)), r) for r in rows]
    scored = [(s, r) for s, r in scored if s is not None]
    best_score, best_row = max(scored, key=lambda t: t[0]) if scored else (None, rows[-1])
    final_row = rows[-1]

    return {
        "best_epoch": col(best_row, *epoch_candidates),
        "best_precision": as_float(col(best_row, *precision_candidates)),
        "best_recall": as_float(col(best_row, *recall_candidates)),
        "best_map50": as_float(col(best_row, *map50_key_candidates)),
        "best_map50_95": best_score,
        "final_map50": as_float(col(final_row, *map50_key_candidates)),
        "final_map50_95": as_float(col(final_row, *map5095_key_candidates)),
    }


def append_run_result(summary_csv_path, row: dict) -> None:
    summary_csv_path = Path(summary_csv_path)
    summary_csv_path.parent.mkdir(parents=True, exist_ok=True)
    is_new = not summary_csv_path.exists()

    full_row = {k: row.get(k, "") for k in FIELDNAMES}

    with open(summary_csv_path, "a", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=FIELDNAMES)
        if is_new:
            writer.writeheader()
        writer.writerow(full_row)


def build_excel_summary(summary_csv_path, out_xlsx_path):
    """가능하면 진짜 .xlsx를, openpyxl이 없으면 csv 그대로를 반환한다."""
    summary_csv_path = Path(summary_csv_path)
    out_xlsx_path = Path(out_xlsx_path)

    try:
        import openpyxl
        from openpyxl.styles import Font
    except ImportError:
        return summary_csv_path

    with open(summary_csv_path, "r", encoding="utf-8-sig") as f:
        rows = list(csv.reader(f))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "sweep_summary"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, value in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    if rows:
        for cell in ws[1]:
            cell.font = Font(bold=True)
        ws.freeze_panes = "A2"
        for col_cells in ws.columns:
            length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(max(length + 2, 10), 40)

    out_xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx_path)
    return out_xlsx_path
